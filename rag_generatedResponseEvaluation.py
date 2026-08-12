import os
import torch
import pandas as pd
import re
from pathlib import Path
from tqdm import tqdm
from transformers import BitsAndBytesConfig
from sentence_transformers import SentenceTransformer, util
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# LlamaIndex Imports
from llama_index.core import VectorStoreIndex, SimpleDirectoryReader, Settings
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.postprocessor import SimilarityPostprocessor
from llama_index.llms.huggingface import HuggingFaceLLM
from llama_index.embeddings.huggingface import HuggingFaceEmbedding

# =============================================
# 🔹 1. SCRIPT CONFIGURATION
# =============================================
# --- Security ---
# NOTE: Set your Hugging Face token as an environment variable
# e.g., export HUGGINGFACE_TOKEN='hf_YOUR_TOKEN_HERE'

# --- RAG Paths ---
PDF_PATH = Path(r"C:\Users\Gx\Desktop\RAG\1_data\report_separated_gpt4_response_v2_gpt_generalisation.pdf")
INPUT_CSV_PATH = Path(
    r"C:\Users\Gx\Desktop\FineTune\GYData\question_sets_llama_generalisation\question_set_d_test_rag.csv")

# --- Output Paths ---
# All output files will be based on this base path
OUTPUT_DIR = Path("generalisation_eva/bert2_gpt")
BASE_FILENAME = "gpt4_knowledge_d_test_rag"

# --- RAG Models ---
EMBED_MODEL_RAG = "allenai/scibert_scivocab_uncased"  # For retrieving documents
LLAMA_MODEL_RAG = "meta-llama/Meta-Llama-3-8B-Instruct"

# --- Evaluation Model ---
EMBED_MODEL_EVAL = 'all-MiniLM-L6-v2'  # For scoring similarity
# intfloat/e5-large-v2
# BAAI/bge-small-en-v1.5
# sentence-transformers/all-mpnet-base-v2
# allenai/scibert_scivocab_uncased

# --- Evaluation Settings ---
THRESHOLD = 0.7  # Scores below this will be highlighted
HIGHLIGHT_COLOR = "FFFF9999"  # Light red fill (aRGB format)

# --- General Settings ---
MAX_ROWS = None  # Set to None to process all rows


# =============================================
# 🔹 2. PART 1: RAG PIPELINE FUNCTIONS
# =============================================

def initialize_components_rag():
    """Initialize RAG models with optimized settings"""
    print("Initializing RAG components (Llama-3 and E5)...")
    quantization_config = BitsAndBytesConfig(
        load_in_4bit=True,
        bnb_4bit_compute_dtype=torch.float16,
        bnb_4bit_quant_type="nf4",
    )

    llm = HuggingFaceLLM(
        model_name=LLAMA_MODEL_RAG,
        tokenizer_name=LLAMA_MODEL_RAG,
        system_prompt="You are a bridge engineering assistant. You are provided with historical bridge inspection records. Recommend a repair proposal based on these historical records.",
        context_window=8192,
        max_new_tokens=256,
        device_map="auto",
        model_kwargs={
            "quantization_config": quantization_config,
            "torch_dtype": torch.float16,
            "trust_remote_code": True
        },
        generate_kwargs={
            "temperature": 0.3,
            "do_sample": False,
        }
    )

    embed_model = HuggingFaceEmbedding(
        model_name=EMBED_MODEL_RAG,
        device="cuda" if torch.cuda.is_available() else "cpu",
        cache_folder="./.cache"
    )

    Settings.node_parser = SentenceSplitter(
        chunk_size=300,
        chunk_overlap=50,
        paragraph_separator="\nFinding ",
        secondary_chunking_regex=r"Finding \d+:"
    )
    Settings.embed_model = embed_model
    Settings.llm = llm

    return llm, embed_model


def load_pdf_documents():
    """Load PDF with validation"""
    print(f"Loading PDF from: {PDF_PATH}")
    if not PDF_PATH.exists():
        raise FileNotFoundError(f"PDF not found at {PDF_PATH}")
    return SimpleDirectoryReader(input_files=[str(PDF_PATH)]).load_data()


def build_rag_pipeline():
    """Build optimized RAG pipeline"""
    documents = load_pdf_documents()
    index = VectorStoreIndex.from_documents(documents)
    return index.as_query_engine(
        similarity_top_k=4,
        node_postprocessors=[SimilarityPostprocessor(similarity_cutoff=0.7)],
        response_mode="compact"
    )


def generate_repair_solution(query, query_engine):
    """Generate standardized repair solution"""
    prompt = f"""
    DEFECT: {query}
    REQUIREMENTS:
    1. Provide ONLY the technical repair method
    2. Maximum 1 sentence
    3. No explanations or disclaimers
    4. If matching a Finding, use format: "Finding [X]: [solution]"
    """
    try:
        response = query_engine.query(prompt)
        return str(response).strip()
    except Exception as e:
        print(f"Error processing query: {str(e)}")
        return "Error: Could not generate solution"


def clean_response(response_text):
    """Enforce standardized output format"""
    clean_text = re.split(r'(Note:|However|Analysis:|Please)', response_text)[0]
    sentences = re.findall(r'(.*?[.!?])', clean_text.strip())
    solution = sentences[0] if sentences else clean_text.strip()
    if "Finding" in solution:
        solution = re.sub(r'Finding (\d+)[:\-]?\s*', r'Finding \1: ', solution)
    return solution


def run_rag_processing(input_path, query_engine, max_rows=None):
    """Process CSV, generate RAG responses, and return DataFrame"""
    print(f"Loading input CSV: {input_path}")
    df = pd.read_csv(input_path)
    if max_rows:
        df = df.head(max_rows)

    if 'Cluster_ID' not in df.columns:
        df['Cluster_ID'] = ""

    df['Standardized_Response'] = ""

    for idx, row in tqdm(df.iterrows(), total=len(df), desc="PART 1: Processing RAG"):
        query = str(row['UserInput']).strip()
        if not query:
            df.at[idx, 'Standardized_Response'] = "Empty query"
            continue
        try:
            raw_response = generate_repair_solution(query, query_engine)
            final_response = clean_response(raw_response)
            df.at[idx, 'Standardized_Response'] = final_response
        except Exception as e:
            print(f"Failed on row {idx}: {str(e)}")
            df.at[idx, 'Standardized_Response'] = f"Error: {str(e)}"

    print("RAG processing complete.")
    return df


# =============================================
# 🔹 3. PART 2: EVALUATION FUNCTIONS
# =============================================

def initialize_eval_model():
    """Initialize the Sentence Transformer for scoring"""
    print(f"Initializing evaluation model ({EMBED_MODEL_EVAL})...")
    device = "cuda" if torch.cuda.is_available() else "cpu"
    return SentenceTransformer(EMBED_MODEL_EVAL).to(device)


def calculate_cosine_similarity(text1, text2, model, device):
    """Calculate cosine similarity between two text strings"""
    try:
        ref_embedding = model.encode(text1, convert_to_tensor=True, device=device)
        gen_embedding = model.encode(text2, convert_to_tensor=True, device=device)
        return util.cos_sim(ref_embedding, gen_embedding).item()
    except Exception as e:
        print(f"Error in similarity calculation: {e}")
        return 0.0


def evaluate_responses(rag_df, eval_model):
    """Evaluate RAG responses against reference responses"""
    device = "cuda" if torch.cuda.is_available() else "cpu"
    results = []

    # Use column names from the RAG output DataFrame
    for _, row in tqdm(rag_df.iterrows(), total=len(rag_df), desc="PART 2: Evaluating Responses"):
        similarity = calculate_cosine_similarity(
            row["Response"],  # The ground truth
            row["Standardized_Response"],  # The RAG output
            eval_model,
            device
        )

        results.append({
            "cluster_id": row["Cluster_ID"],
            "user_input": row["UserInput"],
            "reference_response": row["Response"],
            "generated_response": row["Standardized_Response"],
            "cosine_similarity": similarity,
            "above_threshold": similarity >= THRESHOLD
        })

    return pd.DataFrame(results)


def analyze_clusters(results_df):
    """Calculate robustness metrics for each cluster"""
    cluster_metrics = []
    for cluster_id, group in results_df.groupby('cluster_id'):
        if cluster_id == "":  # Skip rows with no cluster ID
            continue
        cluster_size = len(group)
        similarities = group['cosine_similarity']
        metrics = {
            'cluster_id': cluster_id,
            'cluster_size': cluster_size,
            'consistency_score': (group['above_threshold'].mean()) * 100,
            'avg_similarity': similarities.mean(),
            'std_deviation': similarities.std(),
            'min_score': similarities.min(),
            'max_score': similarities.max(),
            'example_question': group['user_input'].iloc[0],
            'example_response': group['reference_response'].iloc[0]
        }
        cluster_metrics.append(metrics)
    return pd.DataFrame(cluster_metrics)


def calculate_overall_metrics(cluster_df, results_df):
    """Calculate overall evaluation metrics"""
    return {
        'mean_consistency': cluster_df['consistency_score'].mean() if not cluster_df.empty else 0,
        'mean_std_dev': cluster_df['std_deviation'].mean() if not cluster_df.empty else 0,
        'overall_accuracy': (results_df['above_threshold'].mean()) * 100,
        'total_clusters': len(cluster_df),
        'total_queries': len(results_df)
    }


def save_results(results_df, cluster_df, overall_metrics, base_name, output_dir):
    """Save all results to files"""
    print(f"Saving results to {output_dir} with base name {base_name}...")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Define file paths
    scores_csv = output_dir / f"{base_name}_scores.csv"
    clusters_csv = output_dir / f"{base_name}_clusters.csv"
    highlighted_xlsx = output_dir / f"{base_name}_highlighted.xlsx"

    # Save detailed results
    results_df.to_csv(scores_csv, index=False)
    print(f"Saved detailed scores to: {scores_csv}")

    # Save cluster metrics
    if not cluster_df.empty:
        cluster_df.to_csv(clusters_csv, index=False)
        print(f"Saved cluster metrics to: {clusters_csv}")

    # Save highlighted Excel file
    save_highlighted_excel(results_df, overall_metrics, highlighted_xlsx)
    print(f"Saved highlighted Excel to: {highlighted_xlsx}")


def save_highlighted_excel(df, metrics, xlsx_path):
    """Save Excel file with low scores highlighted"""
    df.to_excel(xlsx_path, index=False)
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Find the cosine_similarity column
    sim_col_idx = -1
    for col in ws.iter_cols(min_row=1, max_row=1):
        if col[0].value == "cosine_similarity":
            sim_col_idx = col[0].column

    # Highlight low scores
    if sim_col_idx != -1:
        highlight_fill = PatternFill(start_color=HIGHLIGHT_COLOR, fill_type="solid")
        for row in ws.iter_rows(min_row=2, min_col=sim_col_idx, max_col=sim_col_idx, max_row=ws.max_row):
            try:
                if float(row[0].value) < THRESHOLD:
                    row[0].fill = highlight_fill
            except (ValueError, TypeError):
                continue

    # Add summary
    summary_row = ws.max_row + 2
    ws.cell(summary_row, 1, "SUMMARY METRICS").font = Font(bold=True)
    ws.cell(summary_row + 1, 1, "Overall Accuracy")
    ws.cell(summary_row + 1, 2, f"{metrics['overall_accuracy']:.2f}%")
    ws.cell(summary_row + 2, 1, "Mean Consistency")
    ws.cell(summary_row + 2, 2, f"{metrics['mean_consistency']:.2f}%")
    ws.cell(summary_row + 3, 1, "Mean Std Dev")
    ws.cell(summary_row + 3, 2, f"{metrics['mean_std_dev']:.4f}")

    wb.save(xlsx_path)


# =============================================
# 🔹 4. MAIN EXECUTION FLOW
# =============================================
if __name__ == "__main__":
    try:
        # --- PART 1: RUN RAG PIPELINE ---
        print("=" * 60)
        print("PART 1: STARTING RAG PIPELINE")
        print("=" * 60)
        llm, embed_model_rag = initialize_components_rag()
        query_engine = build_rag_pipeline()
        rag_results_df = run_rag_processing(
            input_path=INPUT_CSV_PATH,
            query_engine=query_engine,
            max_rows=MAX_ROWS
        )

        # --- PART 2: RUN EVALUATION ---
        print("\n" + "=" * 60)
        print("PART 2: STARTING EVALUATION PIPELINE")
        print("=" * 60)
        eval_model = initialize_eval_model()
        eval_results_df = evaluate_responses(rag_results_df, eval_model)
        cluster_df = analyze_clusters(eval_results_df)
        overall_metrics = calculate_overall_metrics(cluster_df, eval_results_df)

        # --- PART 3: SUMMARY & SAVE ---
        print("\n" + "=" * 50)
        print("OVERALL EVALUATION SUMMARY")
        print("=" * 50)
        print(f"Average Cosine Similarity: {eval_results_df['cosine_similarity'].mean():.4f}")
        print(f"Overall Accuracy (>{THRESHOLD}): {overall_metrics['overall_accuracy']:.2f}%")
        print(f"Mean Cluster Consistency: {overall_metrics['mean_consistency']:.2f}%")
        print(f"Mean Cluster Std Dev: {overall_metrics['mean_std_dev']:.4f}")
        print(f"Total Clusters Analyzed: {overall_metrics['total_clusters']}")
        print(f"Total Queries Processed: {overall_metrics['total_queries']}")
        print("=" * 50)

        save_results(
            eval_results_df,
            cluster_df,
            overall_metrics,
            BASE_FILENAME,
            OUTPUT_DIR
        )

        print("\n✅ Combined RAG and Evaluation complete!")

    except Exception as e:
        print(f"\n❌ Fatal error: {str(e)}")