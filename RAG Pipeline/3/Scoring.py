# #
# # import pandas as pd
# # from datasets import Dataset
# # from sentence_transformers import SentenceTransformer, util
# # from tqdm import tqdm
# # import torch
# #
# # # 🔹 Paths and Settings
# # test_csv_file = r"..\2_rag\defect_report_separated_llama3_response_counted_v3_question_set_a_technical_accuracy.csv"  # Replace with your CSV file path (e.g., "bridge_inspection_dialogues.csv")
# # output_csv_file = r"score_GPTLlama_separated_gpt4_response_v2_llama3_response_counted_v3.csv"  # Output file for scores (changed name for clarity)
# # device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
# #
# # # 🔹 Load Sentence Transformer for Cosine Similarity
# # sentence_transformer = SentenceTransformer('all-MiniLM-L6-v2').to(device)  # Lightweight model for embeddings
# #
# # # 🔹 Load and Prepare Test Data
# # try:
# #     df = pd.read_csv(test_csv_file)
# # except FileNotFoundError:
# #     print(f"Error: File {test_csv_file} not found. Please check the path.")
# #     exit(1)
# # except pd.errors.EmptyDataError:
# #     print("Error: The CSV file is empty.")
# #     exit(1)
# #
# # # Verify required columns exist
# # required_columns = ["UserInput", "Response", "Standardized_Response"]
# # if not all(col in df.columns for col in required_columns):
# #     print(f"Error: CSV must contain {required_columns} columns.")
# #     exit(1)
# #
# # # Drop rows with missing values in required columns
# # df = df.dropna(subset=required_columns)
# #
# # # df = df.head(3)  # Uncomment to limit evaluation for testing
# # # print(f"Limited evaluation to the first 5 rows of {test_csv_file}")
# #
# # test_dataset = Dataset.from_pandas(df)
# #
# # # 🔹 Scoring Function (Cosine Similarity Only)
# # def evaluate_responses_cosine(dataset):
# #     cosine_sim_scores = []
# #
# #     for example in tqdm(dataset, desc="Evaluating Responses"):
# #         reference_response = example["Response"]
# #         generated_response = example["Standardized_Response"]
# #
# #         # Cosine Similarity
# #         ref_embedding = sentence_transformer.encode(reference_response, convert_to_tensor=True, device=device)
# #         gen_embedding = sentence_transformer.encode(generated_response, convert_to_tensor=True, device=device)
# #         cosine_sim = util.cos_sim(ref_embedding, gen_embedding).item()
# #         cosine_sim_scores.append(cosine_sim)
# #
# #         # Print scores for each row
# #         print(f"\nRow Evaluation:")
# #         print(f"User Input: {example['UserInput']}")
# #         print(f"Reference Response: {reference_response}")
# #         print(f"Generated Response: {generated_response}")
# #         print(f"Cosine Similarity: {cosine_sim:.4f}")
# #
# #     # Calculate average
# #     avg_cosine_sim = sum(cosine_sim_scores) / len(cosine_sim_scores) if cosine_sim_scores else 0
# #
# #     return {
# #         "cosine_sim_scores": cosine_sim_scores,
# #         "avg_cosine_sim": avg_cosine_sim
# #     }
# #
# # # 🔹 Run Evaluation
# # print("Starting evaluation of responses...")
# # results = evaluate_responses_cosine(test_dataset)
# #
# # # 🔹 Add Scores to DataFrame
# # df["Cosine_Similarity"] = results["cosine_sim_scores"]
# #
# # # 🔹 Print Average Results
# # print("\nAverage Evaluation Results:")
# # print(f"Average Cosine Similarity: {results['avg_cosine_sim']:.4f}")
# # print(f"Number of Samples Evaluated: {len(df)}")
# #
# # # 🔹 Save Results to CSV
# # df.to_csv(output_csv_file, index=False)
# # print(f"Scored results saved to {output_csv_file}")
#
# import pandas as pd
# from datasets import Dataset
# from sentence_transformers import SentenceTransformer, util
# from tqdm import tqdm
# import torch
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill
#
# # 🔹 Configurable Parameters
# THRESHOLD = 0.7  # Scores below this will be highlighted (modify this value as needed)
# HIGHLIGHT_COLOR = "FFFF9999"  # Light red fill (aRGB format - first FF is alpha channel)
#
# # 🔹 Paths and Settings
# test_csv_file = r"..\2_rag\defect_report_separated_llama3_response_counted_v2_question_set_a_robustness_e5.csv"
# output_csv_file = r"score_llama3_separated_llama_response_v2_llama3_response_counted_v3_e5.csv"
# output_xlsx_file = r"score_llama3_separated_llama_response_v2_llama3_response_counted_v3_highlighted_e5.xlsx"
# device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
#
# # 🔹 Load Sentence Transformer for Cosine Similarity
# sentence_transformer = SentenceTransformer('all-MiniLM-L6-v2').to(device)
#
# # 🔹 Load and Prepare Test Data
# try:
#     df = pd.read_csv(test_csv_file)
# except FileNotFoundError:
#     print(f"Error: File {test_csv_file} not found. Please check the path.")
#     exit(1)
# except pd.errors.EmptyDataError:
#     print("Error: The CSV file is empty.")
#     exit(1)
#
# # Verify required columns exist
# required_columns = ["UserInput", "Response", "Standardized_Response"]
# if not all(col in df.columns for col in required_columns):
#     print(f"Error: CSV must contain {required_columns} columns.")
#     exit(1)
#
# # Drop rows with missing values in required columns
# df = df.dropna(subset=required_columns)
# test_dataset = Dataset.from_pandas(df)
#
#
# # 🔹 Scoring Function (Cosine Similarity Only)
# def evaluate_responses_cosine(dataset):
#     cosine_sim_scores = []
#
#     for example in tqdm(dataset, desc="Evaluating Responses"):
#         reference_response = example["Response"]
#         generated_response = example["Standardized_Response"]
#
#         # Cosine Similarity
#         ref_embedding = sentence_transformer.encode(reference_response, convert_to_tensor=True, device=device)
#         gen_embedding = sentence_transformer.encode(generated_response, convert_to_tensor=True, device=device)
#         cosine_sim = util.cos_sim(ref_embedding, gen_embedding).item()
#         cosine_sim_scores.append(cosine_sim)
#
#         print(f"\nRow Evaluation:")
#         print(f"User Input: {example['UserInput']}")
#         print(f"Reference Response: {reference_response}")
#         print(f"Generated Response: {generated_response}")
#         print(f"Cosine Similarity: {cosine_sim:.4f}")
#
#     avg_cosine_sim = sum(cosine_sim_scores) / len(cosine_sim_scores) if cosine_sim_scores else 0
#     return {
#         "cosine_sim_scores": cosine_sim_scores,
#         "avg_cosine_sim": avg_cosine_sim
#     }
#
#
# # 🔹 Run Evaluation
# print("Starting evaluation of responses...")
# results = evaluate_responses_cosine(test_dataset)
#
# # 🔹 Add Scores to DataFrame
# df["Cosine_Similarity"] = results["cosine_sim_scores"]
#
#
# # 🔹 Calculate percentage above threshold
# def calculate_metrics(df, threshold):
#     total = len(df)
#     above_threshold = len(df[df["Cosine_Similarity"] >= threshold])
#     percentage_above = (above_threshold / total) * 100 if total > 0 else 0
#     return percentage_above
#
#
# percentage_above = calculate_metrics(df, THRESHOLD)
#
# # 🔹 Add summary metrics to DataFrame
# summary_row = {
#     "UserInput": "SUMMARY METRICS",
#     "Response": "",
#     "Standardized_Response": "",
#     "Cosine_Similarity": f"Threshold: {THRESHOLD}",
#     "Percentage_Above_Threshold": f"{percentage_above:.2f}%",
#     "Average_Score": f"{results['avg_cosine_sim']:.4f}"
# }
#
# # Append summary row to DataFrame
# df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)
#
# # 🔹 Print Average Results
# print("\nAverage Evaluation Results:")
# print(f"Average Cosine Similarity: {results['avg_cosine_sim']:.4f}")
# print(f"Percentage Above Threshold ({THRESHOLD}): {percentage_above:.2f}%")
# print(f"Number of Samples Evaluated: {len(df) - 1}")  # Subtract 1 for summary row
#
# # 🔹 Save Results to CSV
# df.to_csv(output_csv_file, index=False)
# print(f"\nScored results saved to {output_csv_file}")
#
#
# # 🔹 Save to Excel with Highlighting
# def highlight_low_scores(df, threshold=THRESHOLD, color=HIGHLIGHT_COLOR):
#     """Save DataFrame to Excel with cells below threshold highlighted"""
#     # Make a copy to avoid modifying the original DataFrame
#     df_excel = df.copy()
#
#     # Remove summary row before saving to Excel (we'll add it back later)
#     summary_data = df_excel[df_excel["UserInput"] == "SUMMARY METRICS"]
#     df_excel = df_excel[df_excel["UserInput"] != "SUMMARY METRICS"]
#
#     # First save to Excel
#     df_excel.to_excel(output_xlsx_file, index=False)
#
#     # Now open the file and apply formatting
#     wb = load_workbook(output_xlsx_file)
#     ws = wb.active
#
#     # Create highlight style with proper aRGB format
#     highlight_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
#
#     # Find the column with cosine similarity scores
#     for col_idx, header in enumerate(ws.iter_cols(values_only=True), 1):
#         if header[0] == "Cosine_Similarity":
#             for row in range(2, ws.max_row + 1):  # Skip header row
#                 cell = ws.cell(row=row, column=col_idx)
#                 try:
#                     if float(cell.value) < threshold:
#                         cell.fill = highlight_fill
#                 except (ValueError, TypeError):
#                     continue
#             break
#
#     # Add summary row at the bottom
#     summary_row_num = ws.max_row + 2
#     ws.cell(row=summary_row_num, column=1, value="SUMMARY METRICS")
#     ws.cell(row=summary_row_num, column=4, value=f"Threshold: {threshold}")
#     ws.cell(row=summary_row_num, column=5, value=f"{percentage_above:.2f}%")
#     ws.cell(row=summary_row_num, column=6, value=f"{results['avg_cosine_sim']:.4f}")
#
#     # Bold the summary row
#     for col in range(1, 7):
#         ws.cell(row=summary_row_num, column=col).font = ws.cell(row=summary_row_num, column=col).font.copy(bold=True)
#
#     # Save the formatted file
#     wb.save(output_xlsx_file)
#     print(f"Highlighted Excel file saved to {output_xlsx_file}")
#     print(f"Cells with scores below {threshold} are highlighted")
#
#
# # Save highlighted version
# highlight_low_scores(df)
import pandas as pd
from datasets import Dataset
from sentence_transformers import SentenceTransformer, util
from tqdm import tqdm
import torch
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import numpy as np
import os

# =============================================
# 🔹 CONFIGURATION SECTION (Edit these as needed)
# =============================================
THRESHOLD = 0.7  # Scores below this will be highlighted
HIGHLIGHT_COLOR = "FFFF9999"  # Light red fill (aRGB format)

test_csv_file = r"..\2_rag\generalisation_eva\defect_report_separated_gpt4_response_counted_v2_question_set_d_test_rag_UAE.csv"
output_dir = r"..\2_rag\robustness_eva"

# Create output directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Output files (using your exact naming convention)
output_csv_file = r"..\2_rag\robustness_eva\score_llama3_separated_llama_response_v2_llama3_response_counted_v3_mpnet.csv"
output_xlsx_file = r"..\2_rag\robustness_eva\score_llama3_separated_llama_response_v2_llama3_response_counted_v3_highlighted_mpnet.xlsx"
cluster_metrics_file = r"..\2_rag\robustness_eva\cluster_robustness_metrics_mpnet.csv"

# =============================================
# 🔹 MODEL INITIALIZATION
# =============================================
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
sentence_transformer = SentenceTransformer('all-MiniLM-L6-v2').to(device)


# =============================================
# 🔹 DATA LOADING AND VALIDATION
# =============================================
def load_and_validate_data(csv_path):
    """Load CSV data and validate structure"""
    try:
        df = pd.read_csv(csv_path)
    except FileNotFoundError:
        print(f"Error: File {csv_path} not found.")
        exit(1)
    except pd.errors.EmptyDataError:
        print("Error: CSV file is empty.")
        exit(1)

    # Expected columns (case-insensitive check)
    expected_columns = {
        'cluster_id': ['cluster_id', 'clusterid', 'cluster'],
        'userinput': ['userinput', 'user_input', 'question'],
        'response': ['response', 'answer', 'reference'],
        'standardized_response': ['standardized_response', 'generated', 'llm_output']
    }

    # Find matching columns
    column_map = {}
    for standard_name, variants in expected_columns.items():
        for col in df.columns:
            if col.lower() in variants:
                column_map[standard_name] = col
                break

    # Verify we found all required columns
    missing = [name for name in expected_columns if name not in column_map]
    if missing:
        print(f"Error: Missing required columns - {missing}")
        print(f"Available columns: {list(df.columns)}")
        exit(1)

    # Rename columns to standard names
    df = df.rename(columns={v: k for k, v in column_map.items()})

    # Drop rows with missing values
    df = df.dropna(subset=list(expected_columns.keys()))

    return Dataset.from_pandas(df)


# =============================================
# 🔹 EVALUATION FUNCTIONS
# =============================================
def calculate_cosine_similarity(reference, generated):
    """Calculate cosine similarity between two text responses"""
    ref_embedding = sentence_transformer.encode(reference, convert_to_tensor=True, device=device)
    gen_embedding = sentence_transformer.encode(generated, convert_to_tensor=True, device=device)
    return util.cos_sim(ref_embedding, gen_embedding).item()


def evaluate_responses(dataset):
    """Evaluate all responses in the dataset"""
    results = []

    for example in tqdm(dataset, desc="Evaluating Responses"):
        similarity = calculate_cosine_similarity(
            example["response"],
            example["standardized_response"]
        )

        results.append({
            "cluster_id": example["cluster_id"],
            "user_input": example["userinput"],
            "reference_response": example["response"],
            "generated_response": example["standardized_response"],
            "cosine_similarity": similarity,
            "above_threshold": similarity >= THRESHOLD
        })

    return pd.DataFrame(results)


# =============================================
# 🔹 CLUSTER ANALYSIS FUNCTIONS
# =============================================
def analyze_clusters(results_df):
    """Calculate robustness metrics for each cluster"""
    cluster_metrics = []

    for cluster_id, group in results_df.groupby('cluster_id'):
        cluster_size = len(group)
        similarities = group['cosine_similarity']

        metrics = {
            'cluster_id': cluster_id,
            'cluster_size': cluster_size,
            'consistency_score': (group['above_threshold'].mean()) * 100,
            'avg_similarity': similarities.mean(),
            'std_deviation': similarities.std(),
            'min_score': similarities.min(),
            'max_score': similarities.max(),
            'example_question': group['user_input'].iloc[0],
            'example_response': group['reference_response'].iloc[0]
        }
        cluster_metrics.append(metrics)

    return pd.DataFrame(cluster_metrics)


def calculate_overall_metrics(cluster_df, results_df):
    """Calculate overall evaluation metrics"""
    return {
        'mean_consistency': cluster_df['consistency_score'].mean(),
        'mean_std_dev': cluster_df['std_deviation'].mean(),
        'overall_accuracy': (results_df['above_threshold'].mean()) * 100,
        'total_clusters': len(cluster_df),
        'total_queries': len(results_df)
    }


# =============================================
# 🔹 OUTPUT FUNCTIONS
# =============================================
def save_results(results_df, cluster_df, overall_metrics):
    """Save all results to files"""
    # Save detailed results
    results_df.to_csv(output_csv_file, index=False)

    # Save cluster metrics
    cluster_df.to_csv(cluster_metrics_file, index=False)

    # Save highlighted Excel file
    save_highlighted_excel(results_df, overall_metrics)

    print(f"\nResults saved to: {output_dir}")


def save_highlighted_excel(df, metrics):
    """Save Excel file with low scores highlighted"""
    df.to_excel(output_xlsx_file, index=False)
    wb = load_workbook(output_xlsx_file)
    ws = wb.active

    # Highlight low scores
    highlight_fill = PatternFill(start_color=HIGHLIGHT_COLOR, fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_col=4, max_row=ws.max_row):
        try:
            if float(row[3].value) < THRESHOLD:  # cosine_similarity column
                row[3].fill = highlight_fill
        except (ValueError, TypeError):
            continue

    # Add summary
    summary_row = ws.max_row + 2
    ws.cell(summary_row, 1, "SUMMARY METRICS").font = Font(bold=True)
    ws.cell(summary_row, 2, f"Overall Accuracy: {metrics['overall_accuracy']:.2f}%")
    ws.cell(summary_row, 3, f"Mean Consistency: {metrics['mean_consistency']:.2f}%")
    ws.cell(summary_row, 4, f"Mean Std Dev: {metrics['mean_std_dev']:.4f}")

    wb.save(output_xlsx_file)


# =============================================
# 🔹 MAIN EXECUTION FLOW
# =============================================
if __name__ == "__main__":
    print("Starting evaluation process...")

    # 1. Load and validate data
    print("\nLoading and validating data...")
    test_dataset = load_and_validate_data(test_csv_file)

    # 2. Evaluate responses
    print("\nEvaluating responses...")
    results_df = evaluate_responses(test_dataset)

    # 3. Analyze clusters
    print("\nAnalyzing clusters...")
    cluster_df = analyze_clusters(results_df)
    overall_metrics = calculate_overall_metrics(cluster_df, results_df)

    # 4. Print summary
    print("\n" + "=" * 50)
    print("EVALUATION SUMMARY")
    print("=" * 50)
    print(f"Average Cosine Similarity: {results_df['cosine_similarity'].mean():.4f}")
    print(f"Overall Accuracy: {overall_metrics['overall_accuracy']:.2f}%")
    print(f"Mean Standard Deviation: {overall_metrics['mean_std_dev']:.4f}")
    print(f"Total Clusters: {overall_metrics['total_clusters']}")
    print(f"Total Queries: {overall_metrics['total_queries']}")

    # 5. Save results
    print("\nSaving results...")
    save_results(results_df, cluster_df, overall_metrics)

    print("\nEvaluation complete!")